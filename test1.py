import openpyxl

# 打开现有的Excel文件
workbook = openpyxl.load_workbook('C:/Users/pqs-xsz-xia/Desktop/test.xlsx')
# 选择要操作的工作表
worksheet = workbook.active
# 获取现有内容的最后一行
last_row = worksheet.max_row
# 要写入的新内容
new_content = 'New Content'
# 将新内容写入下一个单元格
next_cell = worksheet.cell(row=last_row+1, column=1)
cell_value = worksheet.cell(row=last_row, column=1).value
print(cell_value)
next_cell.value = new_content
# 保存修改后的Excel文件
workbook.save('C:/Users/pqs-xsz-xia/Desktop/test.xlsx')
