import requests
import re
import openpyxl
from bs4 import BeautifulSoup


def get_page_source(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        response.encoding = 'utf-8'  # 指定为UTF-8编码方式
        return response.text
    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")
        return None


def extract_content(page_source):
    link = []
    pattern = r'<li><div class="imgr".*?<a href="(.*?)".*?>(.*?)</a>.*?(\d{4}-\d{2}-\d{2})</span>.*?</li>'
    matches = re.findall(pattern, page_source, re.DOTALL)
    count = 0
    for match in matches:
        link.append(match[0])
    return link


for j in range(27, 40):
    # 指定要爬取的网站 URL
    url = f"https://www.mkaq.org/sggl/shigukb/index_{j}.shtml"
    # 调用函数获取页面源代码
    page_source = get_page_source(url)
    # 调用函数提取内容
    link = extract_content(page_source)
    for i in link:
        print(i)
        html = get_page_source(i)
        soup = BeautifulSoup(html, 'html.parser')
        div = soup.find('div', class_='article_content')
        text = div.get_text().strip()
        # 打开现有的Excel文件
        workbook = openpyxl.load_workbook('C:/Users/pqs-xsz-xia/Desktop/test.xlsx')
        # 选择要操作的工作表
        worksheet = workbook.active
        # 获取现有内容的最后一行
        last_row = worksheet.max_row
        # 要写入的新内容
        new_content = text
        # 将新内容写入下一个单元格
        next_cell = worksheet.cell(row=last_row + 1, column=1)
        cell_value = worksheet.cell(row=last_row, column=1).value

        if cell_value != text:
            next_cell.value = new_content
            # 保存修改后的Excel文件
            workbook.save('C:/Users/pqs-xsz-xia/Desktop/test.xlsx')
        else:
            print('无需更新')
